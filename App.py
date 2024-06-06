import tkinter
import customtkinter as ctk
import os
import tkinter as tk
import pandas as pd
import pkg_resources
from glob import glob
from openpyxl import load_workbook
from CTkMessagebox import CTkMessagebox
from cachetools import cached, TTLCache
from PIL import ImageTk, Image
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import Rule
from tkinter import ttk

# Define a cache with a maximum size of 100 entries and a time-to-live of 300 seconds (5 minutes)
cache = TTLCache(maxsize=100, ttl=300)


class App(ctk.CTk):
    def __init__(self, conn, db2_conn):
        super().__init__(fg_color="#1A2932")
        self.geometry("1500x725")
        self.title("Load Automation Process Check")
        self.stage_row_df = None
        self.client_data = None
        self.conn = conn
        self.db2_conn = db2_conn
        self.setup_ui()

        # Load the green image and resize them
        green_image_path = pkg_resources.resource_filename(__name__, 'counts_match.png')
        green_image = Image.open(green_image_path)
        green_image_resized = green_image.resize((40, 40), Image.Resampling.LANCZOS)
        self.green_image = ImageTk.PhotoImage(green_image_resized)
        self.green_image_label = ctk.CTkLabel(self, image=self.green_image, text="",
                                              width=40, height=40, fg_color="#0a545c")

        # Load the red image and resize them
        red_image_path = pkg_resources.resource_filename(__name__, 'not_match.png')
        red_image = Image.open(red_image_path)
        red_image_resized = red_image.resize((40, 40), Image.Resampling.LANCZOS)
        self.red_image = ImageTk.PhotoImage(red_image_resized)
        self.red_image_label = ctk.CTkLabel(self, image=self.red_image, text="",
                                            width=40, height=40, fg_color="#0a545c")

        # Set the appearance
        ctk.set_appearance_mode("dark")
        # Set the appearance
        theme_path = pkg_resources.resource_filename(__name__, 'gainwellthemeV2.json')
        ctk.set_default_color_theme(theme_path)

    def setup_ui(self):
        self.header_setup()
        self.button_setup()
        self.batch_setup()
        self.stage_setup()
        self.update_spreadsheet_process()
        self.refresh_app()

    def header_setup(self):
        # HEADER FRAME
        self.header_frame = ctk.CTkFrame(self, width=1450, height=50, fg_color="#0a545c")
        self.header_frame.place(x=25, y=25)

        # HEADER LABEL
        self.header_label = ctk.CTkLabel(self.header_frame, text="LOAD AUTOMATION PROCESS CHECK",
                                         width=1350, font=ctk.CTkFont(size=15, weight="bold"),
                                         text_color="White", justify="center")
        self.header_label.place(x=0, y=10)

    def button_setup(self):
        # BUTTON FRAME
        self.button_frame = ctk.CTkFrame(self, width=1050, height=125, fg_color="#233845")
        self.button_frame.place(x=25, y=85)

        # CODE SELECTION RADIO BUTTONS
        self.radio_var = tkinter.IntVar(value=0)
        self.parent_code_select = ctk.CTkRadioButton(self.button_frame, text="Parent Code",
                                                     command=self.handle_radio_selection,
                                                     font=ctk.CTkFont(size=13), variable=self.radio_var, value=1)
        self.parent_code_select.place(x=50, y=20)
        self.client_code_select = ctk.CTkRadioButton(self.button_frame, text="Client Code",
                                                     command=self.handle_radio_selection,
                                                     font=ctk.CTkFont(size=13), variable=self.radio_var, value=2)
        self.client_code_select.place(x=50, y=70)

        # PARENT OR CLIENT CODE ENTRY
        self.code_entry = ctk.CTkEntry(self.button_frame, placeholder_text="Enter Code",
                                       placeholder_text_color="white", justify="center",
                                       font=ctk.CTkFont(size=15), width=150, height=40,
                                       fg_color="#5C6870", border_width=0, corner_radius=0)
        self.code_entry.place(x=200, y=20)

        # ENTER BUTTON
        self.enter_button = ctk.CTkButton(self.button_frame, text="Enter", command=self.handle_enter,
                                          font=ctk.CTkFont(size=15, weight="bold"), width=150, height=35,
                                          fg_color="#00EFAE", text_color="#233845")
        self.enter_button.place(x=200, y=70)

    def batch_setup(self):
        # BATCH TABLE LABEL & BUTTON
        self.batch_label = ctk.CTkLabel(self.button_frame, text="Batch Table",
                                        font=ctk.CTkFont(size=15),
                                        text_color="White", justify="center", )
        self.batch_label.place(x=530, y=25)

        self.batch_button = ctk.CTkButton(self.button_frame, text="Run", command=self.run_batch_query,
                                          font=ctk.CTkFont(size=15, weight="bold"), width=150, height=35,
                                          fg_color="#00EFAE", text_color="#233845")
        self.batch_button.place(x=500, y=70)

        # BATCH TABLE LISTBOX LABEL FRAME
        self.batch_label_frame = ctk.CTkFrame(self, width=1050, height=40, fg_color="#0a545c")
        self.batch_label_frame.place(x=25, y=220)

        # BATCH LABEL
        self.batch_frame_label = ctk.CTkLabel(self.batch_label_frame, width=1050, height=20,
                                              text="Batch Table Results",
                                              font=ctk.CTkFont(size=14, weight="bold"), justify="center")
        self.batch_frame_label.place(x=0, y=10)

        # BATCH TABLE FRAME
        self.batch_table_frame = ctk.CTkFrame(self, width=1050, height=175, fg_color="#233845")
        self.batch_table_frame.place(x=25, y=265)

        # Create a custom style
        self.style = ttk.Style()
        self.style.theme_use("default")

        # Configure the style to set background and foreground colors for Treeview
        self.style.configure("Treeview", background="#5C6870", foreground="#FFFFFF")

        # Configure the style to set background and foreground colors for Treeview headings
        self.style.configure("Treeview.Heading", background="#233845", foreground="white")

        # BATCH COLUMN NAMES
        batch_column_names = ["PARENT_CODE", "CLIENT_CODE", "BATCH_ID", "RUN_STATUS_CD", "FILE_NAME", "SRC_FILE_COUNT",
                              "REJECT_COUNT", "FUSP_UPDATE"]

        # Initialize Treeview
        self.batch_tree = ttk.Treeview(self.batch_table_frame, columns=batch_column_names, show="headings")
        self.batch_tree.place(x=5, y=5, width=1040, height=160)

        # Add column headings
        for col in batch_column_names:
            self.batch_tree.heading(col, text=col)

        # BATCH COLUMN NAMES
        batch_column_names = ["PARENT_CODE", "CLIENT_CODE", "BATCH_ID", "RUN_STATUS_CD", "FILE_NAME", "SRC_FILE_COUNT",
                              "REJECT_COUNT", "FUSP_UPDATE"]

        # Initialize Treeview
        self.batch_tree = ttk.Treeview(self.batch_table_frame, columns=batch_column_names, show="headings")
        self.batch_tree.place(x=5, y=5, width=1040, height=160)

        # Add column headings
        for col in batch_column_names:
            self.batch_tree.heading(col, text=col)

        # Configure Treeview's column width
        for col in batch_column_names:
            if col == "FILE_NAME":
                # Set a wider width (e.g., 200) for the FILE_NAME column
                self.batch_tree.column(col, width=200, anchor=tk.CENTER)
            elif col == 'RUN_STATUS_CD':
                self.batch_tree.column(col, width=50, anchor=tk.CENTER)
            else:
                # Set a standard width (e.g., 50) for other columns
                self.batch_tree.column(col, width=30, anchor=tk.CENTER)

        # Bind selection event to the batch treeview
        self.batch_tree.bind("<<TreeviewSelect>>", self.on_batch_select)

    def stage_setup(self):
        # STAGE TABLE LABEL & BUTTON
        self.stage_label = ctk.CTkLabel(self.button_frame, text="Stage Table",
                                        font=ctk.CTkFont(size=15),
                                        text_color="White", justify="center")
        self.stage_label.place(x=780, y=25)

        self.stage_button = ctk.CTkButton(self.button_frame, text="Run", command=self.run_stage_query,
                                          font=ctk.CTkFont(size=15, weight="bold"), width=150, height=35,
                                          fg_color="#00EFAE", text_color="#233845")
        self.stage_button.place(x=750, y=70)

        # STAGE TABLE LISTBOX LABEL FRAME
        self.stage_label_frame = ctk.CTkFrame(self, width=1050, height=40, fg_color="#0a545c")
        self.stage_label_frame.place(x=25, y=450)

        # STAGE LABEL
        self.stage_frame_label = ctk.CTkLabel(self.stage_label_frame, width=1050, height=20,
                                              text="Stage Table Results",
                                              font=ctk.CTkFont(size=14, weight="bold"), justify="center")
        self.stage_frame_label.place(x=0, y=10)

        # STAGE TABLE FRAME
        self.stage_table_frame = ctk.CTkFrame(self, width=1050, height=175, fg_color="#233845")
        self.stage_table_frame.place(x=25, y=495)

        # Create a custom style
        self.style = ttk.Style()
        self.style.theme_use("default")

        # Configure the style to set background and foreground colors for Treeview
        self.style.configure("Treeview", background="#5C6870", foreground="#FFFFFF")

        # Configure the style to set background and foreground colors for Treeview headings
        self.style.configure("Treeview.Heading", background="#233845", foreground="white")

        # STAGE COLUMN NAMES
        stage_column_names = ["PARENT_CODE", "CLIENT_CODE", "BATCH_ID", "TOTAL_STAGE", "SRC_STATUS", "APPROVAL_STATUS",
                              "POSTING_ID"]

        # Initialize Treeview
        self.stage_tree = ttk.Treeview(self.stage_table_frame, columns=stage_column_names, show="headings")
        self.stage_tree.place(x=5, y=5, width=1040, height=160)

        # Add column headings
        for col in stage_column_names:
            self.stage_tree.heading(col, text=col)

        # Configure Treeview's column width
        for col in stage_column_names:
            self.stage_tree.column(col, width=100, anchor=tk.CENTER)

        # Bind selection event to the stage treeview
        self.stage_tree.bind("<<TreeviewSelect>>", self.on_stage_select)

    def update_spreadsheet_process(self):
        # UPDATE LABEL FRAME
        self.update_spreadsheet_frame = ctk.CTkFrame(self, width=390, height=40, fg_color="#0a545c")
        self.update_spreadsheet_frame.place(x=1085, y=85)

        # UPDATE LABEL
        self.update_label = ctk.CTkLabel(self.update_spreadsheet_frame, text="Update Data",
                                         font=ctk.CTkFont(size=14, weight="bold"), width=390,
                                         text_color="White", justify="center")
        self.update_label.place(x=0, y=5)

        # DATA LABEL FRAME
        self.data_frame = ctk.CTkFrame(self, width=390, height=535, fg_color="#233845")
        self.data_frame.place(x=1085, y=135)

        # Batch Listboxes
        self.listboxes_parent_code = {}
        self.listboxes_client_code = {}
        self.listboxes_batch_id = {}
        self.listboxes_run_status_cd = {}
        self.listboxes_file_name = {}
        self.listboxes_src_file_count = {}
        self.listboxes_reject_count = {}
        self.listboxes_fusp_update = {}

        # Stage Listboxes
        self.listboxes_total_stage = {}
        self.listboxes_src_status = {}
        self.listboxes_approval_status = {}
        self.listboxes_posting_id = {}

        listbox_info = [
            ("PARENT_CODE", 20, self.listboxes_parent_code),
            ("CLIENT_CODE", 65, self.listboxes_client_code),
            ("BATCH_ID", 110, self.listboxes_batch_id),
            ("RUN_STATUS_CD", 155, self.listboxes_run_status_cd),
            ("FILE_NAME", 160, self.listboxes_file_name),
            ("SRC_FILE_COUNT", 205, self.listboxes_src_file_count),
            ("REJECT_COUNT", 250, self.listboxes_reject_count),
            ("FUSP_UPDATE", 295, self.listboxes_fusp_update),
            ("TOTAL_STAGE", 340, self.listboxes_total_stage),
            ("SRC_STATUS", 385, self.listboxes_src_status),
            ("APPROVAL_STATUS", 430, self.listboxes_approval_status),
            ("POSTING_ID", 475, self.listboxes_posting_id)
        ]

        # Loop through each listbox info and create the label and listbox
        y_position = 2
        for label_text, y_position, listbox_dict in listbox_info:
            # Create and place label
            label = ctk.CTkLabel(self.data_frame, width=150, height=30, text=label_text, bg_color="#1A2932",
                                 font=ctk.CTkFont(size=14, weight="bold"))
            label.place(x=10, y=y_position)

            # Create and place listbox
            listbox = tk.Listbox(self.data_frame, height=1, width=175 // 7, justify="left",
                                 font=ctk.CTkFont(size=12, weight="bold"), bg='#5C6870',
                                 bd=2, fg='white', highlightbackground='#5C6870', highlightthickness=3,
                                 relief="flat", highlightcolor='#96A3B0')
            listbox.place(x=165, y=y_position)

            # Save the listbox reference using the label as key
            listbox_dict[label_text] = listbox

        # LOAD DATA BUTTON
        self.load_button = ctk.CTkButton(self, text="Load Data", command=self.load_data,
                                         font=ctk.CTkFont(size=15, weight="bold"), width=150, height=35,
                                         fg_color="#00EFAE", text_color="#233845")
        self.load_button.place(x=1085, y=680)

        # UPDATE SPREADSHEET BUTTON
        self.update_button = ctk.CTkButton(self, text="Export", command=self.update_report,
                                           font=ctk.CTkFont(size=15, weight="bold"), width=150, height=35,
                                           fg_color="#00EFAE", text_color="#233845")
        self.update_button.place(x=1325, y=680)

        # REFRESH BUTTON
        self.refresh_button = ctk.CTkButton(self, text="Refresh", command=self.refresh_app,
                                            font=ctk.CTkFont(size=15, weight="bold"), width=150, height=35,
                                            fg_color="#00EFAE", text_color="#233845")
        self.refresh_button.place(x=50, y=680)

        self.counts_label = ctk.CTkLabel(self, text="",
                                         font=ctk.CTkFont(size=18, weight="bold"), width=300,
                                         text_color="White", justify="center")
        self.counts_label.place(x=450, y=680)

    def check_counts_match(self):
        batch_src_file_count = self.listboxes_src_file_count["SRC_FILE_COUNT"].get(0)
        batch_reject_count = self.listboxes_reject_count["REJECT_COUNT"].get(0)
        batch_fusp_update = self.listboxes_fusp_update["FUSP_UPDATE"].get(0)

        stage_total_stage = self.listboxes_total_stage["TOTAL_STAGE"].get(0)
        stage_src_status = self.listboxes_src_status["SRC_STATUS"].get(0)
        stage_approval_status = self.listboxes_approval_status["APPROVAL_STATUS"].get(0)
        stage_posting_id = self.listboxes_posting_id["POSTING_ID"].get(0)

        if (batch_src_file_count == stage_total_stage and
                batch_reject_count == stage_src_status and
                batch_fusp_update == stage_approval_status and
                batch_src_file_count == stage_posting_id):
            self.counts_label.configure(text="Counts Match!", text_color="#00b050")
            self.green_image_label.place(x=1400, y=85)
            self.red_image_label.place_forget()  # Remove the red image label if it's visible
        else:
            self.counts_label.configure(text="Counts Do Not Match!", text_color="#ff0000")
            self.red_image_label.place(x=1400, y=85)
            self.green_image_label.place_forget()

    def refresh_app(self):
        # Clear all data and reset UI components to their original state
        self.code_entry.delete(0, tk.END)  # Clear client code entry
        # Clear listbox data
        for listbox_dict in [self.listboxes_parent_code, self.listboxes_client_code, self.listboxes_batch_id,
                             self.listboxes_run_status_cd, self.listboxes_file_name, self.listboxes_src_file_count,
                             self.listboxes_reject_count, self.listboxes_fusp_update, self.listboxes_total_stage,
                             self.listboxes_src_status, self.listboxes_approval_status, self.listboxes_posting_id]:
            for listbox in listbox_dict.values():
                listbox.delete(0, tk.END)
        # Clear Treeview data
        self.batch_tree.delete(*self.batch_tree.get_children())
        self.stage_tree.delete(*self.stage_tree.get_children())

        # Reset the counts_label
        self.counts_label.configure(text="", text_color="black")

    def get_code(self):
        return self.code_entry.get()

    def handle_radio_selection(self):
        try:
            if self.radio_var.get() == 1:  # If "Parent Code" radio button is selected
                self.selected_code_type = "Parent Code"  # Save the code type selection
            elif self.radio_var.get() == 2:  # If "Client Code" radio button is selected
                self.selected_code_type = "Client Code"  # Save the code type selection
            else:
                # Handle the case where no radio button is selected
                CTkMessagebox(title="Error", message="Please select a code type.")
        except Exception as e:
            # Handle exceptions
            CTkMessagebox(title="Error", message=f"Error occurred: {e}")

    def handle_enter(self):
        code = None
        try:
            code = self.get_code()  # To get the code entered by the user
            if not code:
                CTkMessagebox(title="Error", message="Please enter a code.")
                return

            self.selected_code = code  # Save the entered code
            CTkMessagebox(title="Attention!", message="Code has been entered! Press the Run button for Batch Table.")
        except Exception as e:
            CTkMessagebox(title="Error", message=f"Error occurred: {e}")

    def run_batch_query(self):
        try:
            if not self.selected_code_type:
                CTkMessagebox(title="Error", message="Please enter a code and select code type.")
                return

            # Construct the query based on the selected code type
            if self.selected_code_type == "Parent Code":
                query = """
                    SELECT 
                        PARENT_CD, 
                        PARENT_CD AS CLIENT_CD, 
                        BATCH_ID, 
                        RUN_STATUS_CD, 
                        FILE_NAME, 
                        SUM(SRC_FILE_COUNT) AS SRC_FILE_COUNT, 
                        SUM(REJECT_COUNT) AS REJECT_COUNT, 
                        SUM(FUSP_UPDATE_COUNT) AS FUSP_UPDATE
                    FROM COV.CAV_LOAD_RPT_BATCH
                    WHERE PARENT_CD = ?
                    GROUP BY PARENT_CD, BATCH_ID, RUN_STATUS_CD, FILE_NAME
                    ORDER BY BATCH_ID DESC
                    LIMIT 10;
                """
            elif self.selected_code_type == "Client Code":
                query = """
                    SELECT PARENT_CD, CLIENT_CD, BATCH_ID, RUN_STATUS_CD, FILE_NAME, SRC_FILE_COUNT, REJECT_COUNT, FUSP_UPDATE_COUNT AS FUSP_UPDATE
                    FROM COV.CAV_LOAD_RPT_BATCH
                    WHERE CLIENT_CD = ?
                    ORDER BY BATCH_ID DESC
                    LIMIT 10;
                """

            # Call batch_query with the constructed query and the retrieved code
            result = self.batch_query(query, self.selected_code)

            # Display the batch results
            if result:
                self.display_batch_results(result)
                CTkMessagebox(title="Attention", message="Select a batch row!")
        except Exception as e:
            # Handle exceptions
            CTkMessagebox(title="Error", message=f"Error occurred during batch query: {e}")

    def display_batch_results(self, results):
        # Clear existing data in the treeview
        self.batch_tree.delete(*self.batch_tree.get_children())

        # Format and insert new data into the treeview
        for row in results:
            formatted_row = [str(item) for item in row]  # Convert all items to strings

            # Check if RUN_STATUS_CD is "LRA_FAILED"
            if formatted_row[2] == "LRA_FAILED":
                # Highlight the row in red
                self.batch_tree.insert("", tk.END, values=formatted_row, tags=("red_row",))
            else:
                # Insert normally
                self.batch_tree.insert("", tk.END, values=formatted_row)

        # Apply tag configuration to set the background color of rows with the "red_row" tag to red
        self.batch_tree.tag_configure("red_row", background="red")

    def display_stage_results(self, results):
        # Clear existing data in the treeview
        self.stage_tree.delete(*self.stage_tree.get_children())

        # Format and insert new data into the treeview
        for row in results:
            formatted_row = [str(item) for item in row]  # Convert all items to strings
            self.stage_tree.insert("", tk.END, values=formatted_row)

    def on_batch_select(self, event):
        # Get selected item
        selected_item = self.batch_tree.focus()

        # Extract data from selected item
        selected_data = self.batch_tree.item(selected_item, "values")

        # Create a DataFrame with the selected row data
        batch_row_df = pd.DataFrame({
            self.batch_tree['columns'][0]: [selected_data[0]],
            self.batch_tree['columns'][1]: [selected_data[1]],
            self.batch_tree['columns'][2]: [selected_data[2]],
            self.batch_tree['columns'][3]: [selected_data[3]],
            self.batch_tree['columns'][4]: [selected_data[4]],
            self.batch_tree['columns'][5]: [selected_data[5]],
            self.batch_tree['columns'][6]: [selected_data[6]],
            self.batch_tree['columns'][7]: [selected_data[7]]
        })

        # Now you can store this DataFrame for later use
        self.batch_row_df = batch_row_df

        # Extract batch_id from the selected data
        batch_id = selected_data[2]

        # Store batch_id for later use
        self.batch_id = batch_id

        # Show message box
        CTkMessagebox(title="Attention", message="Run Stage Table Process")

    def on_stage_select(self, event):
        # Get selected item
        selected_item = self.stage_tree.focus()

        # Extract data from selected item
        selected_data = self.stage_tree.item(selected_item, "values")

        # Create a DataFrame with the selected row data
        stage_row_df = pd.DataFrame({
            self.stage_tree['columns'][0]: [selected_data[0]],
            self.stage_tree['columns'][1]: [selected_data[1]],
            self.stage_tree['columns'][2]: [selected_data[2]],
            self.stage_tree['columns'][3]: [selected_data[3]],
            self.stage_tree['columns'][4]: [selected_data[4]],
            self.stage_tree['columns'][5]: [selected_data[5]],
            self.stage_tree['columns'][6]: [selected_data[6]]
        })

        # Now you can store this DataFrame for later use
        self.stage_row_df = stage_row_df

        # Extract total_stage from selected data
        total_stage = selected_data[3]

        # store total_stage for later use
        self.total_stage = total_stage

        # Show message box
        CTkMessagebox(title="Attention", message="Load Data to Listboxes")

    def run_stage_query(self):
        try:
            # Construct the query based on the selected radio button
            if self.radio_var.get() == 1:  # If "Parent Code" radio button is selected
                query = """
            WITH valid_parents AS (
                SELECT 
                    BATCH_ID, 
                    MAX(PARENT_CD) AS PARENT_CD 
                FROM 
                    COV.CAV_LOAD_RPT_BATCH 
                WHERE 
                    PARENT_CD IS NOT NULL
                GROUP BY 
                    BATCH_ID
            )
            SELECT
                COALESCE(B.PARENT_CD, VP.PARENT_CD) AS PARENT_CD,
                COALESCE(B.PARENT_CD, VP.PARENT_CD) AS CLIENT_CD,
                S.BATCH_ID AS BATCH_ID,
                COUNT(*) AS TOTAL_STAGE,
                SUM(CASE WHEN S.SRC_STATUS LIKE 'R' THEN 1 ELSE 0 END) AS SRC_STATUS,
                SUM(CASE WHEN S.APPROVAL_STATUS LIKE 'R' THEN 1 ELSE 0 END) AS APPROVAL_STATUS,
                SUM(CASE WHEN S.POSTING_ID IS NOT NULL THEN 1 ELSE 0 END) AS POSTING_ID
            FROM COV.CAV_LOAD_RPT_STAGE AS S
            LEFT JOIN COV.CAV_LOAD_RPT_BATCH AS B
                ON S.CLIENT_CD = B.CLIENT_CD
                AND S.BATCH_ID = B.BATCH_ID
            LEFT JOIN valid_parents AS VP
                ON S.BATCH_ID = VP.BATCH_ID
            WHERE S.BATCH_ID = ?
            GROUP BY
                COALESCE(B.PARENT_CD, VP.PARENT_CD),
                S.BATCH_ID;
                """
            elif self.radio_var.get() == 2:  # If "Client Code" radio button is selected
                query = """
                    SELECT
                        CASE WHEN B.PARENT_CD IS NULL THEN S.CLIENT_CD ELSE B.PARENT_CD END AS PARENT_CD,
                        S.CLIENT_CD AS CLIENT_CD,
                        S.BATCH_ID AS BATCH_ID,
                        COUNT(*) AS TOTAL_STAGE,
                        SUM(CASE WHEN S.SRC_STATUS LIKE 'R' THEN 1 ELSE 0 END) AS SRC_STATUS,
                        SUM(CASE WHEN S.APPROVAL_STATUS LIKE 'R' THEN 1 ELSE 0 END) AS APPROVAL_STATUS,
                        SUM(CASE WHEN S.POSTING_ID IS NOT NULL THEN 1 ELSE 0 END) AS POSTING_ID
                    FROM COV.CAV_LOAD_RPT_STAGE AS S
                    LEFT JOIN COV.CAV_LOAD_RPT_BATCH AS B
                        ON S.CLIENT_CD = B.CLIENT_CD
                        AND S.BATCH_ID = B.BATCH_ID
                    WHERE S.BATCH_ID = ?
                    GROUP BY
                        CASE WHEN B.PARENT_CD IS NULL THEN S.CLIENT_CD ELSE B.PARENT_CD END,
                        S.CLIENT_CD,
                        S.BATCH_ID;
                """

            # Call stage_query with the constructed query and the retrieved batch ID
            self.stage_query(query, self.batch_id)

            CTkMessagebox(title="Attention", message="Select a Stage row!")
        except Exception as e:
            # Handle exceptions
            print(f"Error occurred during stage query: {e}")
            CTkMessagebox(title="Error", message=f"Error occurred during stage query: {e}")

    @cached(cache)
    def batch_query(self, query, code):
        try:
            cursor = self.db2_conn.cursor()
            cursor.execute(query, (code,))
            result = cursor.fetchall()

            return result  # Return the fetched data
        except Exception as e:
            # Handle exceptions, such as database connection errors
            CTkMessagebox(title="Error", message=f"Error occurred during batch query: {e}")

    @cached(cache)
    def stage_query(self, query, batch_id):
        try:
            cursor = self.db2_conn.cursor()
            cursor.execute(query, (batch_id))
            result = cursor.fetchall()

            if result:
                self.display_stage_results(result)
        except Exception as e:
            # Handle exceptions, such as database connection errors
            CTkMessagebox(title="Error", message=f"Error occurred during stage query: {e}")

    # In get_processed_data method
    def get_processed_data(self):
        # Create a dictionary to hold processed data
        processed_data = {
            'parent_code': self.batch_row_df['PARENT_CODE'],
            'client_code': self.batch_row_df['CLIENT_CODE'],
            'batch_id': self.batch_row_df['BATCH_ID'],
            'run_status_cd': self.batch_row_df['RUN_STATUS_CD'],
            'file_name': self.batch_row_df['FILE_NAME'],
            'src_file_count': self.batch_row_df['SRC_FILE_COUNT'],
            'reject_count': self.batch_row_df['REJECT_COUNT'],
            'fusp_update': self.batch_row_df['FUSP_UPDATE'],
            'total_stage': self.stage_row_df['TOTAL_STAGE'],
            'src_status': self.stage_row_df['SRC_STATUS'],
            'approval_status': self.stage_row_df['APPROVAL_STATUS'],
            'posting_id': self.stage_row_df['POSTING_ID'],
        }

        return processed_data

    def get_batch_info(self):
        parent_code = self.listboxes_parent_code["PARENT_CODE"].get(tk.ACTIVE)
        client_code = self.listboxes_client_code["CLIENT_CODE"].get(tk.ACTIVE)
        batch_id = self.listboxes_batch_id["BATCH_ID"].get(tk.ACTIVE)
        file_name = self.listboxes_file_name["FILE_NAME"].get(tk.ACTIVE)
        src_file_count = self.listboxes_src_file_count["SRC_FILE_COUNT"].get(tk.ACTIVE)
        reject_count = self.listboxes_reject_count["REJECT_COUNT"].get(tk.ACTIVE)
        fusp_update = self.listboxes_fusp_update["FUSP_UPDATE"].get(tk.ACTIVE)

        # Return this data as a dictionary with keys matching your Excel file headers
        return {
            'PARENT CODE': parent_code,
            'CLIENT CODE': client_code,
            'BATCH ID': batch_id,
            'FILE NAME': file_name,
            'SRC_FILE_COUNT (BATCH TABLE)': src_file_count,
            'REJECT_COUNT (BATCH TABLE)': reject_count,
            'FUSP_UPDATE_COUNT (BATCH TABLE)': fusp_update
        }

    def get_stage_info(self):
        total_stage = self.listboxes_total_stage["TOTAL_STAGE"].get(tk.ACTIVE)
        src_status = self.listboxes_src_status["SRC_STATUS"].get(tk.ACTIVE)
        approval_status = self.listboxes_approval_status["APPROVAL_STATUS"].get(tk.ACTIVE)
        posting_id = self.listboxes_file_name["FILE_NAME"].get(tk.ACTIVE)

        # Return this data as a dictionary with keys matching your Excel file headers
        return {
            'TOTAL BATCH RECORD COUNT (STAGE TABLE)': total_stage,
            'SRC_STATUS = R (STAGE TABLE)': src_status,
            'APPROVAL STATUS = R (STAGE TABLE)': approval_status,
            'POSTING_ID IS NOT NULL (STAGE TABLE)': posting_id
        }

    # To load data to the list-box's
    def load_data(self):
        processed_data = self.get_processed_data()
        self.populate_data(processed_data)

        # After loading data, check if the counts match
        self.check_counts_match()

    # Populate the data
    def populate_data(self, data):
        # Processing Batch data
        batch_info = data.get('parent_code', pd.Series())
        for value in batch_info:
            self.listboxes_parent_code["PARENT_CODE"].insert('end', value)

        batch_info = data.get('client_code', pd.Series())
        for value in batch_info:
            self.listboxes_client_code["CLIENT_CODE"].insert('end', value)

        batch_info = data.get('batch_id', pd.Series())
        for value in batch_info:
            self.listboxes_batch_id["BATCH_ID"].insert('end', value)

        batch_info = data.get('run_status_cd', pd.Series())
        for value in batch_info:
            self.listboxes_run_status_cd["RUN_STATUS_CD"].insert('end', value)

        batch_info = data.get('file_name', pd.Series())
        for value in batch_info:
            self.listboxes_file_name["FILE_NAME"].insert('end', value)

        batch_info = data.get('src_file_count', pd.Series())
        for value in batch_info:
            self.listboxes_src_file_count["SRC_FILE_COUNT"].insert('end', value)

        batch_info = data.get('reject_count', pd.Series())
        for value in batch_info:
            self.listboxes_reject_count["REJECT_COUNT"].insert('end', value)

        batch_info = data.get('fusp_update', pd.Series())
        for value in batch_info:
            self.listboxes_fusp_update["FUSP_UPDATE"].insert('end', value)

        # Processing Stage data
        stage_info = data.get('total_stage', pd.Series())
        for value in stage_info:
            self.listboxes_total_stage["TOTAL_STAGE"].insert('end', value)

        stage_info = data.get('src_status', pd.Series())
        for value in stage_info:
            self.listboxes_src_status["SRC_STATUS"].insert('end', value)

        stage_info = data.get('approval_status', pd.Series())
        for value in stage_info:
            self.listboxes_approval_status["APPROVAL_STATUS"].insert('end', value)

        stage_info = data.get('posting_id', pd.Series())
        for value in stage_info:
            self.listboxes_posting_id["POSTING_ID"].insert('end', value)

    @cached(cache)
    def update_query(self, code):
        try:
            cursor = self.conn.cursor()
            query = """
                SELECT LR_UPDATE
                FROM DL_CAV_LR.CLIENT_REF
                WHERE CLIENT_CD = ?;
            """
            cursor.execute(query, (code,))
            result = cursor.fetchone()

            return result
        except Exception as e:
            # Handle exceptions, such as database connection errors
            CTkMessagebox(title="Error", message=f"Error occurred during update query: {e}")

    def update_report(self):
        code = self.get_code()
        report_location = self.get_report_location(code)

        # Retrieve the data from the listboxes
        batch_info = {
            'PARENT_CODE': self.listboxes_parent_code["PARENT_CODE"].get('active'),
            'CLIENT_CODE': self.listboxes_client_code["CLIENT_CODE"].get('active'),
            'BATCH_ID': self.listboxes_batch_id["BATCH_ID"].get('active'),
            'FILE_NAME': self.listboxes_file_name["FILE_NAME"].get('active'),
            'SRC_FILE_COUNT': self.listboxes_src_file_count["SRC_FILE_COUNT"].get('active'),
            'REJECT_COUNT': self.listboxes_reject_count["REJECT_COUNT"].get('active'),
            'FUSP_UPDATE': self.listboxes_fusp_update["FUSP_UPDATE"].get('active')
        }
        stage_info = {
            'TOTAL_STAGE': self.listboxes_total_stage["TOTAL_STAGE"].get('active'),
            'SRC_STATUS': self.listboxes_src_status["SRC_STATUS"].get('active'),
            'APPROVAL_STATUS': self.listboxes_approval_status["APPROVAL_STATUS"].get('active'),
            'POSTING_ID': self.listboxes_posting_id["POSTING_ID"].get('active')
        }

        # Create styles for fill colors
        fill_red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        fill_green = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

        # Search for an existing workbook
        existing_file_path = self.search_for_existing_reports(report_location, code)
        if existing_file_path:
            # Load the existing workbook and use its path for saving
            wb = load_workbook(existing_file_path)
            ws = wb['LA Process']
            file_path = existing_file_path
            for row in ws.iter_rows(min_row=1, max_col=12, max_row=ws.max_row):
                cell = row[11]
                if cell.value == 'Y':
                    cell.fill = fill_red
                elif cell.value == 'N':
                    cell.fill = fill_green
        else:
            # Load template and set a new file path if no existing workbook is found
            template_path = 'C:\\Users\\jroynon\\PycharmProjects\\LoadAppTesting\\Load Report Update.xlsx'
            wb = load_workbook(template_path)
            ws = wb['LA Process']
            file_path = os.path.join(report_location, f"{code} - Load Report Update.xlsx")

        # Find the next empty row to avoid overwriting existing data
        next_row = ws.max_row + 1

        # Set font color for text
        blue_font = Font(color="0000FF")  # Blue font color for SRC_FILE_COUNT, REJECT_COUNT, and FUSP_UPDATE
        red_font = Font(color="FF0000")  # Red font color for TOTAL_STAGE, SRC_STATUS, APPROVAL_STATUS, and POSTING_ID

        # Populate the 'Load Reports' sheet with the retrieved data
        for col, col_name in enumerate(['PARENT_CODE', 'CLIENT_CODE', 'BATCH_ID', 'FILE_NAME', 'SRC_FILE_COUNT',
                                        'REJECT_COUNT', 'FUSP_UPDATE', 'TOTAL_STAGE', 'SRC_STATUS', 'APPROVAL_STATUS',
                                        'POSTING_ID'], start=1):
            cell = ws.cell(row=next_row, column=col)
            if 1 <= col <= 11:  # Only apply font color changes to columns A to K
                cell.alignment = Alignment(horizontal='right')  # Ensure right alignment
                if col_name in batch_info:
                    if col_name in ['SRC_FILE_COUNT', 'REJECT_COUNT', 'FUSP_UPDATE']:
                        cell.font = blue_font  # Blue font color for SRC_FILE_COUNT, REJECT_COUNT, and FUSP_UPDATE
                elif col_name in stage_info:
                    if col_name in ['TOTAL_STAGE', 'SRC_STATUS', 'APPROVAL_STATUS', 'POSTING_ID']:
                        cell.font = red_font  # Red font color for TOTAL_STAGE, SRC_STATUS, APPROVAL_STATUS, and POSTING_ID

            cell.value = batch_info.get(col_name, '') if col_name in batch_info else stage_info.get(col_name, '')

        # Add formula to column L
        formula = f'=IF((AND(E{next_row}=H{next_row},F{next_row}=G{next_row},I{next_row}=J{next_row},E{next_row}=K{next_row})),"N","Y")'
        ws[f'L{next_row}'] = formula

        # Save the workbook to the predetermined file path
        try:
            wb.save(file_path)
            # After saving, show a pop-up message indicating success
            workbook_name = os.path.basename(file_path)
            self.open_excel_file(file_path)

            # Correctly format the message to include variables
            CTkMessagebox(title="Workbook Saved", message=f"{workbook_name} has been saved!")
        except Exception as e:
            CTkMessagebox(title="Save Error", message=f"Failed to save the workbook: {e}")

    def search_for_existing_reports(self, directory, code):
        search_patterns = [f"{code} Load Report Update*.xlsx",
                           f"{code} Load Report Reconciliation*.xlsx",
                           f"{code} - Load Report Update*.xlsx",
                           f"{code} - Load Report Reconciliation*.xlsx",
                           f"*Load Report Update*.xlsx",
                           f"*Load Report Reconciliation*.xlsx"
                           ]
        for pattern in search_patterns:
            for filename in glob(os.path.join(directory, pattern)):
                return filename
        return None

    def open_excel_file(self, file_path):
        # Opens an Excel file using the default app
        os.startfile(file_path)

    @cached(cache)
    def get_report_location(self, code):
        try:
            if self.radio_var.get() == 1:
                query = """
                SELECT LR_UPDATE FROM DL_CAV_LR.CLIENT_REF WHERE PARENT_CD = ?
                """
            elif self.radio_var.get() == 2:
                query = """
                SELECT LR_UPDATE FROM DL_CAV_LR.CLIENT_REF WHERE CLIENT_CD = ?
                """
            cursor = self.conn.cursor()
            cursor.execute(query, (code,))
            result = cursor.fetchone()
            return result[0] if result else None
        except Exception as e:
            # Handle exceptions, such as database connection errors
            CTkMessagebox(title="Error", message=f"Error occurred during report location query: {e}")
            return None  # Return None in case of error


if __name__ == "__main__":
    conn = None
    app = App(conn, db2_conn=None)
    app.mainloop()
